import openpyxl
from groq import Groq
from dotenv import load_dotenv
import os
from pathlib import Path
load_dotenv(dotenv_path=Path(__file__).parent / ".env")
import json
from openpyxl.styles import Border, Side, PatternFill

def main():
    transactions = get_transactions()

    raw_entry = []

    for transaction in transactions:
        transaction = transaction["transaction"]
        raw_entry.append(analyze_entries(transaction))

    formatted = []
    for transaction, entry in zip(transactions, raw_entry):
        row1 = {"date": transaction["date"], "particulars": entry["debit_account"] + "   Dr.", "debit":entry[ "amount"], "credit": "" }
        row2 = {"date": "", "particulars": "    To " +  entry["credit_account"] , "debit": "", "credit":entry[ "amount"] }
        row3 = {"date": "", "particulars": entry["narration"], "debit":"", "credit": "" }

        formatted.extend([row1, row2, row3])
    write_journal(formatted)
    print("Journal Saved to Computer!")

def get_transactions():
    workbook = openpyxl.load_workbook(filename ="data.xlsx")
    sheet = workbook.active

    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append({"date": row[0], "transaction": row[1]})

    return data

def analyze_entries(transaction_text):
    client = Groq(api_key=os.getenv("GROQ_API_KEY"))
    response = client.chat.completions.create(
        model="llama-3.3-70b-versatile",
        messages=[
            {"role": "system", "content": """You are a certified chartered accountant in india
Your job is to analyze the given transactions to output a debit account, a credit account, a transaction amount and a narration for the transaction (the narration should always start with the word "Being" and should always be inside ()).
Rules:
analyze each account as:
1. Personal Account. (real person/ artificial person)
2. Real Account. (assets and liabilities)
3. Nominal Account. (Expenses and Losses, Incomes and gains)

For each account these are the rules:
1. Personal Account: Debit the receiver, Credit the giver.
2. Real Account: Debit what comes in, credit what goes out.
3. Nominal Account: Debit all expenses and losses, credit all incomes and gains.

Always respond with ONLY valid JSON in this exact format:
{"debit_account": "<account name>", "credit_account": "<account name>", "amount": <number>, "narration": "<narration>"}

Example:
input: "Mr. Anand started business with cash Rs. 60,000"
output:
{"debit_account": "Cash A/C", "credit_account": "Capital A/C", "amount": 60000, "naration" : "(Being business commences with cash.)"}
"""},

            {"role": "user", "content": transaction_text}      # the actual transaction text
        ],
        response_format={"type": "json_object"}  # keep this if you decide JSON mode is right
    )
    return json.loads(response.choices[0].message.content.strip())

def write_journal(formatted):

    output_workbook = openpyxl.Workbook()
    output_sheet = output_workbook.active
    output_sheet.append(["Date", "Particulars", "Debit Amount (Rs.)", "Credit Amount (Rs.)"])

    header_fill = PatternFill(start_color="FFADD8E6", end_color="FFADD8E6", fill_type="solid")
    for col in range(1, 5):
        cell = output_sheet.cell(row=1, column=col)
        cell.fill = header_fill
        
    total_debit = 0
    total_credit = 0

    for line in formatted:
        row_values = [line["date"], line["particulars"], line["debit"], line["credit"]]
        output_sheet.append(row_values)


        if line["debit"] == "" and line["credit"] == "":
            row_number = output_sheet.max_row
            thin_border = Border(bottom=Side(style="thin"))
            cell = output_sheet.cell(row=row_number, column=2)
            cell.border = thin_border

        if line["debit"] != "":
            total_debit += line["debit"]

        if line["credit"] != "":
            total_credit += line["credit"]


    output_sheet.append(["", "Total", total_debit, total_credit])
    row_number = output_sheet.max_row
    total_border = Border(top=Side(style="thin"), bottom=Side(style="thick"))

    for col in [3, 4]:
        cell = output_sheet.cell(row=row_number, column=col)
        cell.border = total_border

    output_workbook.save("output.xlsx")

main()


